from django.contrib.auth.models import Permission
from django.shortcuts import render

# Create your views here.
from django.shortcuts import render

# Create your views here.

from django.conf import settings
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from .models import LearningPath, LearningTask, Assignment, Result, LearningComplex, LearningComplexPath, AssignmentRepeat
from .filters import LearningPathFilter, AssignmentFilter, ResultFilter, LearningComplexFilter
from .forms import LearningPathForm, LearningTaskForm, AssignmentForm, LearningComplexForm, LearningComplexPathForm, AssignmentRepeatForm
import pandas as pd
from django.shortcuts import render, redirect, reverse, get_object_or_404
from django.http import HttpResponse, JsonResponse
from guardian.mixins import PermissionListMixin
from guardian.mixins import PermissionRequiredMixin
from django.contrib.auth.mixins import PermissionRequiredMixin as GPermissionRequiredMixin
from materials.models import Material
from courses.models import Course
from testing.models import Test
from events.models import Event
from core.models import Subdivision, Employee, Placement
from django.db.models import OuterRef, Subquery, BooleanField, Count, IntegerField
from django.db.models.functions import Coalesce
from django.db.models import F
from django.db.models import Case, When, Value
from django.db import models
from django.db.models import Q
from django.http import HttpResponseNotFound, HttpResponseForbidden, JsonResponse
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from datetime import datetime
from django.db.models import Sum
from django.contrib.auth.decorators import permission_required, login_required
from django.forms import modelformset_factory
from django.forms.formsets import ORDERING_FIELD_NAME
from django import forms
from core.mixins import PreviousPageGetMixinL1, PreviousPageSetMixinL1, \
    PreviousPageGetMixinL2, PreviousPageSetMixinL2, \
    PreviousPageGetMixinL3, PreviousPageSetMixinL3, \
    PreviousPageGetMixinL4, PreviousPageSetMixinL4, \
    PreviousPageGetMixinL5, PreviousPageSetMixinL5, \
    PreviousPageGetMixinL0, PreviousPageSetMixinL0
from reviews.models import Review
from core.models import EmployeesGroupObjectPermission, EmployeesObjectPermission
from reviews.filters import ObjectsReviewFilter
from django.core.paginator import Paginator
from django.contrib.contenttypes.models import ContentType
from core.filters import EmployeesGroupObjectPermissionGroupsFilter, EmployeesObjectPermissionEmployeesFilter
from datetime import timedelta
from datetime import date
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required

# Импортируем логи
import logging
# Создаем логгер с именем 'project'
logger = logging.getLogger('project')

# Список траекторий.
class LearningComplexsView(LoginRequiredMixin, PreviousPageSetMixinL5, PermissionListMixin, ListView):
    # Права доступа
    permission_required = 'learning_path.view_learningcomplex'
    # Модель.
    model = LearningComplex
    # Поле сортировки.
    ordering = 'name'
    # Шаблон.
    template_name = 'learning_complexs.html'
    # Количество объектов на странице
    paginate_by = 6

    # Переопределяем выборку вью.
    def get_queryset(self):
        # Забираем изначальную выборку вью.
        queryset = super().get_queryset()
        # Добавляем модель фильтрации в выборку вью.
        self.filterset = LearningComplexFilter(self.request.GET, queryset, request=self.request)
        # Возвращаем вью новую выборку.
        return self.filterset.qs

    # Добавляем во вью переменные.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)
        # Добавляем фильтрсет.
        context['filterset'] = self.filterset
        # Возвращаем новый набор переменных.
        return context

# Объект категории.
class LearningComplexView(LoginRequiredMixin, PreviousPageGetMixinL5, PreviousPageSetMixinL4, PermissionRequiredMixin, ListView):
    # Права доступа.
    permission_required = 'learning_path.view_learningcomplex'
    accept_global_perms = True
    # Модель.
    model = LearningComplexPath
    # Шаблон.
    template_name = 'learning_complex.html'
    # Количество объектов на странице
    paginate_by = 6

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_complex = LearningComplex.objects.get(pk=self.kwargs.get('pk'))
        return learning_complex

    # Переопределяем выборку вью.
    def get_queryset(self):
        # Переопределяем изначальную выборку.
        learning_complex = self.get_permission_object()
        queryset = super().get_queryset().filter(learning_complex=learning_complex).order_by('position')
        self.qs_count = len(queryset)
        # Возвращаем вью новую выборку.
        return queryset

    # Добавляем во вью переменные.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)
        learning_complex = self.get_permission_object()
        context['object'] = learning_complex
        context['qs_count'] = self.qs_count

        # Получаем доп. информацию.
        if learning_complex.learning_complex_paths.exists():
            # Создаем переменную.
            learning_complex_paths_exists = True
            context['learning_complex_paths_exists']=learning_complex_paths_exists

        # Возвращаем новый набор переменных.
        return context


# Создание траектории.
class LearningComplexCreateView(LoginRequiredMixin, GPermissionRequiredMixin, CreateView):
    # Права доступа.
    permission_required = 'learning_path.add_larningcomplex'
    # Форма.
    form_class = LearningComplexForm
    # Модель.
    model = LearningComplex
    # Шаблон.
    template_name = 'learning_complex_edit.html'

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_complex', kwargs={'pk': self.object.pk})

# Изменение траектории.
class LearningComplexUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    # Права доступа.
    permission_required = 'learning_path.change_learningcomplex'
    # Форма.
    form_class = LearningComplexForm
    accept_global_perms = True
    # Модель.
    model = LearningComplex
    # Шаблон.
    template_name = 'learning_complex_edit.html'

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_complex', kwargs={'pk': self.object.pk})

# Удаление траектории.
class LearningComplexDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    # Права доступа.
    permission_required = 'learning_path.delete_learningcomplex'
    # Модель.
    model = LearningComplex
    accept_global_perms = True
    # Шаблон.
    template_name = 'learning_complex_delete.html'

    # Перенаправление после валидации формы.PermissionListMixin
    def get_success_url(self):
        # Направляем по адресу объектов.
        return reverse('learning_path:learning_complexs')

# Список траекторий программы.
class LearningComplexPathsView(LoginRequiredMixin, PreviousPageSetMixinL3, PermissionRequiredMixin, ListView):
    # Права доступа.
    permission_required = 'learning_path.view_learningcomplex'
    accept_global_perms = True
    # Модель.
    model = LearningComplexPath
    # Поле сортировки.
    ordering = 'position'
    # Шаблон.
    template_name = 'learning_complex_paths.html'
    # Количество объектов на странице.
    paginate_by = 3

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_complex = LearningComplex.objects.get(pk=self.kwargs.get('pk'))
        return learning_complex

    # Переопределяем выборку вью.
    def get_queryset(self):
        # Переопределяем изначальную выборку.
        queryset = super().get_queryset().filter(learning_complex__pk=self.kwargs.get('pk')).order_by('position')
        # Возвращаем вью новую выборку.
        return queryset

    # Переопределеяем переменные вью.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)
        # Забираем путь.
        object = self.get_permission_object()
        # Добавляем во вью.
        context['object'] = object
        # Возвращаем новый набор переменных в контролер.
        return context

# Создание траектории.
class LearningComplexPathCreateView(LoginRequiredMixin, GPermissionRequiredMixin, CreateView):
    # Права доступа.
    permission_required = 'learning_path.add_learningcomplex'
    # Форма.
    form_class = LearningComplexPathForm
    # Модель.
    model = LearningComplexPath
    # Шаблон.
    template_name = 'learning_complex_path_edit.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_complex = LearningComplex.objects.get(pk=self.kwargs.get('pk'))
        return learning_complex

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем позицию.
        learning_complex = self.get_permission_object()
        last_position = learning_complex.learning_complex_paths.order_by('position').values_list('position', flat=True).last()
        if last_position:
            position = last_position + 1
        else:
            position = 1
        initial["position"] = position
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Добавляем программу.
        initial["learning_complex"] = learning_complex
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_complex', kwargs={'pk': self.object.learning_complex.pk})

# Сортировка траекторий комплексной программы.
@login_required
@permission_required('testing.change_learningcomplex')
def learning_paths_ordering(request, pk):

    #Кверисет.
    queryset = LearningComplexPath.objects.filter(learning_complex__pk=pk).order_by('position')

    # Формсет.
    LearningComplexPathSet = modelformset_factory(
        LearningComplexPath,
        fields=('learning_path',),
        can_order=True,
        can_delete=False,
        extra=0,
        widgets={
            'type': forms.HiddenInput(),
        },
    )

    # Если форма отправлена.
    if request.method == 'POST':
        formset = LearningComplexPathSet(request.POST, queryset=queryset)
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data:
                    learning_complex_paths = form.save(commit=False)
                    learning_complex_paths.position = form.cleaned_data[ORDERING_FIELD_NAME]
                    learning_complex_paths.save()
            return redirect('learning_path:learning_complex', pk=pk)

    # Если форма открыта.
    else:
        formset = LearningComplexPathSet(queryset=queryset)

    # Контекст.
    contex = {'formset': formset}

    # Форма.
    return render(request, 'learning_complex_paths_ordering.html', contex)


# Изменение траектории.
class LearningComplexPathUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    # Права доступа.
    permission_required = 'learning_path.change_learningcomplex'
    accept_global_perms = True
    # Форма.
    form_class = LearningComplexPathForm
    # Модель.
    model = LearningComplexPath
    # Шаблон.
    template_name = 'learning_complex_path_edit.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_complex = self.get_object().learning_complex
        return learning_complex

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_complex', kwargs={'pk': self.object.learning_complex.pk})

# Удаление траектории.
class LearningComplexPathDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    # Права доступа.
    permission_required = 'learning_path.delete_learningcomplex'
    accept_global_perms = True
    # Модель.
    model = LearningComplexPath
    # Шаблон.
    template_name = 'learning_complex_path_delete.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_complex = self.get_object().learning_complex
        return learning_complex

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объектов.
        return reverse('learning_path:learning_complex', kwargs={'pk': self.object.learning_complex.pk})

# Список траекторий.
class LearningPathsView(LoginRequiredMixin, PreviousPageSetMixinL3, PermissionListMixin, ListView):
    # Права доступа.
    permission_required = 'learning_path.view_learningpath'
    # Модель.
    model = LearningPath
    # Поле сортировки.
    ordering = 'name'
    # Шаблон.
    template_name = 'learning_paths.html'
    # Количество объектов на странице.
    paginate_by = 6

    # Переопределяем выборку вью.
    def get_queryset(self):
        # Забираем изначальную выборку вью.
        queryset = super().get_queryset()
        # Анотируем кверисет контрольными задачами.
        control_tasks = LearningTask.objects.filter(
            learning_path_id=OuterRef('pk'),
        ).values('learning_path_id').annotate(count=Count('id')).values('count')
        queryset = queryset.annotate(
            control_task_count=Coalesce(Subquery(control_tasks, output_field=IntegerField()), Value(0))
        )
        # Анотируем кверисет результатами.
        latest_result = Result.objects.filter(
            learning_path_id=OuterRef('pk'),
            employee=self.request.user,
            type='learning_path',
        ).order_by('-id').values('status', 'planned_end_date', 'end_date')[:1]
        queryset = queryset.annotate(
            latest_result_status=Coalesce(Subquery(latest_result.values('status')), Value('None')),
            latest_result_planned_end_date=Subquery(latest_result.values('planned_end_date')),
            latest_result_end_date=Subquery(latest_result.values('end_date'))
        )
        queryset = queryset.annotate(
            has_date=Case(
                When(latest_result_planned_end_date__isnull=False, then=Value(1)),
                default=Value(0),
                output_field=models.IntegerField()
            )
        ).order_by('-has_date', 'latest_result_planned_end_date', 'name')
        # Добавляем модель фильтрации в выборку вью.
        self.filterset = LearningPathFilter(self.request.GET, queryset, request=self.request)
        # Возвращаем вью новую выборку.
        return self.filterset.qs

    # Добавляем во вью переменные.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)
        # Добавляем фильтрсет.
        context['filterset'] = self.filterset
        # Возвращаем новый набор переменных.
        return context

# Объект категории.
class LearningPathView(LoginRequiredMixin, PreviousPageGetMixinL3, PreviousPageSetMixinL2, PreviousPageSetMixinL0, PermissionRequiredMixin, ListView):
    # Права доступа.
    permission_required = 'learning_path.view_learningpath'
    accept_global_perms = True
    # Модель.
    model = LearningTask
    # Шаблон.
    template_name = 'learning_path.html'
    # Количество объектов на странице
    paginate_by = 6

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_path = LearningPath.objects.get(pk=self.kwargs.get('pk'))
        return learning_path

    # Переопределяем выборку вью.
    def get_queryset(self):
        # Переопределяем изначальную выборку.
        learning_path = self.get_permission_object()
        queryset = super().get_queryset().filter(learning_path=learning_path).order_by('position')
        self.qs_count = len(queryset)
        # Возвращаем вью новую выборку.
        return queryset

    # Добавляем во вью переменные.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)
        learning_path = self.get_permission_object()
        context['object'] = learning_path
        context['qs_count'] = self.qs_count

        # Получаем доп. информацию.
        if learning_path.learning_tasks.exists():
            # Создаем переменную.
            learning_tasks_exists = True
            context['learning_tasks_exists']=learning_tasks_exists

        # Забираем результат.
        if learning_path.results.filter(employee=self.request.user):
            result = learning_path.results.filter(employee=self.request.user, type='learning_path').latest('id')
            logger.info(f'Результат {result}')
            # Добавляем фильтрсет.
            context['result'] = result

        # Добавляем средний балл.
        # Сначала получаем сумму оценок и количество отзывов.
        reviews_sum = learning_path.reviews.aggregate(Sum('score'))['score__sum']
        reviews_count = learning_path.reviews.count()
        # Вычисляем среднюю оценку.
        if reviews_count > 0:
            reviews_average_mark = round((reviews_sum / reviews_count), 1)
        else:
            reviews_average_mark = 0
        # Добавляем вью.
        context['reviews_average_mark'] = reviews_average_mark

        # Вычисляем число контрольных задач.
        control_task_count = learning_path.learning_tasks.filter(control_task=True).count()
        context['control_task_count'] = control_task_count

        # Добавляем отзывы.
        if Review.objects.filter(type='learning_path', learning_path__pk=self.kwargs.get('pk')).exists():
            reviews_queryset = Review.objects.filter(type='learning_path', learning_path__pk=self.kwargs.get('pk')).order_by('created')
        else:
            reviews_queryset = Review.objects.none()
        context['reviews_qs_count'] = len(reviews_queryset)
        reviews_filter = ObjectsReviewFilter(self.request.GET, queryset=reviews_queryset, request=self.request)
        reviews = reviews_filter.qs
        # Добавляем пагинатор.
        reviews_paginator = Paginator(reviews, 6)
        reviews_page_number = self.request.GET.get('reviews_page')
        reviews_page_obj = reviews_paginator.get_page(reviews_page_number)
        # Добавляем во вью.
        context['reviews_filter'] = reviews_filter
        context['reviews_page_obj'] = reviews_page_obj

        # Забираем отзыв.
        haves_review = False
        if learning_path.reviews.filter(creator=self.request.user).exists():
            haves_review = True
        if settings.DEBUG:
            logger.info(f"Уже есть отзыв: {haves_review}")
        context['haves_review'] = haves_review

        # Добавляем объектные права.
        content_type = ContentType.objects.get_for_model(LearningPath)
        if EmployeesGroupObjectPermission.objects.filter(
            content_type=content_type,
            object_pk=self.kwargs.get('pk')
        ).prefetch_related('content_object').exists():
            group_object_permissions_queryset = EmployeesGroupObjectPermission.objects.filter(
                content_type=content_type,
                object_pk=self.kwargs.get('pk')
            ).prefetch_related('content_object').order_by('group')
        else:
            group_object_permissions_queryset = EmployeesGroupObjectPermission.objects.none()
        context['group_object_permissions_qs_count'] = len(group_object_permissions_queryset)
        group_object_permissions_filter = EmployeesGroupObjectPermissionGroupsFilter(self.request.GET, queryset=group_object_permissions_queryset, request=self.request)
        group_object_permissions = group_object_permissions_filter.qs
        # Добавляем пагинатор.
        group_object_permissions_paginator = Paginator(group_object_permissions, 6)
        group_object_permissions_page_number = self.request.GET.get('group_object_permissions_page')
        group_object_permissions_page_obj = group_object_permissions_paginator.get_page(group_object_permissions_page_number)
        # Добавляем во вью.
        context['group_object_permissions_filter'] = group_object_permissions_filter
        context['group_object_permissions_page_obj'] = group_object_permissions_page_obj

        # Добавляем объектные права.
        if EmployeesObjectPermission.objects.filter(object_pk=self.kwargs.get('pk'), content_type=content_type).prefetch_related('content_object').exists():
            object_permissions_queryset = EmployeesObjectPermission.objects.filter(
                object_pk=self.kwargs.get('pk'),
                content_type=content_type
            ).prefetch_related('content_object').order_by('-id')
        else:
            object_permissions_queryset = EmployeesObjectPermission.objects.none()
        context['object_permissions_qs_count'] = len(object_permissions_queryset)
        object_permissions_filter = EmployeesObjectPermissionEmployeesFilter(self.request.GET, queryset=object_permissions_queryset, request=self.request)
        object_permissions = object_permissions_filter.qs
        # Добавляем пагинатор.
        object_permissions_paginator = Paginator(object_permissions, 6)
        object_permissions_page_number = self.request.GET.get('object_permissions_page')
        object_permissions_page_obj = object_permissions_paginator.get_page(object_permissions_page_number)
        # Добавляем во вью.
        context['object_permissions_filter'] = object_permissions_filter
        context['object_permissions_page_obj'] = object_permissions_page_obj

        # Возвращаем новый набор переменных.
        return context

# Создание траектории.
class LearningPathCreateView(LoginRequiredMixin, GPermissionRequiredMixin, CreateView):
    # Права доступа.
    permission_required = 'learning_path.add_learningpath'
    # Форма.
    form_class = LearningPathForm
    # Модель.
    model = LearningPath
    # Шаблон.
    template_name = 'learning_path_edit.html'

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_path', kwargs={'pk': self.object.pk})

# Изменение траектории.
class LearningPathUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    # Права доступа.
    permission_required = 'learning_path.change_learningpath'
    accept_global_perms = True
    # Форма.
    form_class = LearningPathForm
    # Модель.
    model = LearningPath
    # Шаблон.
    template_name = 'learning_path_edit.html'

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_path', kwargs={'pk': self.object.pk})

# Удаление траектории.
class LearningPathDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    # Права доступа.
    permission_required = 'learning_path.delete_learningpath'
    accept_global_perms = True
    # Модель.
    model = LearningPath
    # Шаблон.
    template_name = 'learning_path_delete.html'

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объектов.
        return reverse('learning_path:learning_paths')

# Список задач.
class LearningPathTasksView(LoginRequiredMixin, PreviousPageSetMixinL1, PermissionRequiredMixin, ListView):
    # Права доступа
    permission_required = 'learning_path.view_learningpath'
    # Модель.
    model = LearningTask
    # Поле сортировки.
    ordering = 'position'
    # Шаблон.
    template_name = 'learning_path_tasks.html'
    # Количество объектов на странице
    paginate_by = 3

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_path = LearningPath.objects.get(pk=self.kwargs.get('pk'))
        return learning_path

    # Переопределяем выборку вью.
    def get_queryset(self):

        # Забираем изначальную выборку вью.
        queryset = super().get_queryset().filter(learning_path__pk=self.kwargs.get('pk')).order_by('position')

        # Создаем подзапросы.
        task_latest_result = Result.objects.filter(
            learning_task_id=OuterRef('pk'), employee=self.request.user
        ).order_by('-id').values('id', 'status', 'learning_path_result__planned_end_date', 'end_date')[:1]
        material_latest_result = Result.objects.filter(
                material=OuterRef('material'),
                employee=self.request.user
            ).order_by('-id').values('status', 'learning_path_result__planned_end_date', 'end_date')[:1]
        course_latest_result = Result.objects.filter(
                course=OuterRef('course'),
                employee=self.request.user
            ).order_by('-id').values('status', 'learning_path_result__planned_end_date', 'end_date')[:1]
        test_latest_result = Result.objects.filter(
                test=OuterRef('test'),
                employee=self.request.user
            ).order_by('-id').values('status', 'learning_path_result__planned_end_date', 'end_date')[:1]

        # Аннотируем queryset
        queryset = queryset.annotate(
            tasks_latest_result_id=Subquery(task_latest_result.values('id')),
            tasks_latest_result_status=Subquery(task_latest_result.values('status')),
            tasks_latest_result_planned_end_date=Subquery(task_latest_result.values('learning_path_result__planned_end_date')),
            tasks_latest_result_end_date=Subquery(task_latest_result.values('end_date')),
            latest_result_status=Case(
                When(material__isnull=False, then=Subquery(material_latest_result.values('status'))),
                When(course__isnull=False, then=Subquery(course_latest_result.values('status'))),
                When(test__isnull=False, then=Subquery(test_latest_result.values('status'))),
                default=Value(None)
            ),
            latest_result_planned_end_date=Case(
                When(material__isnull=False, then=Subquery(material_latest_result.values('learning_path_result__planned_end_date'))),
                When(course__isnull=False, then=Subquery(course_latest_result.values('learning_path_result__planned_end_date'))),
                When(test__isnull=False, then=Subquery(test_latest_result.values('learning_path_result__planned_end_date'))),
                default=Value(None),
            ),
            latest_result_end_date = Case(
                When(material__isnull=False,
                     then=Subquery(material_latest_result.values('end_date'))),
                When(course__isnull=False,
                     then=Subquery(course_latest_result.values('end_date'))),
                When(test__isnull=False,
                     then=Subquery(test_latest_result.values('end_date'))),
                default=Value(None),
            )
        )

        # Добавление блокировки.
        for obj in queryset:
            # Если есть результаты задачи.
            if obj.blocking_tasks:
                # Забираем блокирующие задачи.
                obj.now_blocking_tasks = [
                    blocking_task for blocking_task in obj.blocking_tasks.all()
                    if blocking_task.results.filter(employee=self.request.user).exists() and blocking_task.results.filter(employee=self.request.user).latest('id').status != 'completed'
                ]

        # Возвращаем вью новую выборку.
        return queryset

    # Переопределеяем переменные вью.
    def get_context_data(self, **kwargs):
        # забираем изначальный набор переменных
        context = super().get_context_data(**kwargs)
        # Забираем путь.
        object = LearningPath.objects.get(pk=self.kwargs.get('pk'))
        # Добавляем во вью.
        context['object'] = object
        # Забираем задачи.
        appointed_education_count = self.object_list.filter(
            tasks_latest_result_status='appointed',
        ).count()
        education_in_progress_count = self.object_list.filter(
            tasks_latest_result_status='in_progress',
        ).count()
        completed_education_count = self.object_list.filter(
            tasks_latest_result_status='completed',
        ).count()
        failed_education_count = self.object_list.filter(
            tasks_latest_result_status='failed',
        ).count()
        context['appointed_education_count'] = appointed_education_count
        context['education_in_progress_count'] = education_in_progress_count
        context['completed_education_count'] = completed_education_count
        context['failed_education_count'] = failed_education_count
        # Возвращаем новый набор переменных в контролер.
        return context

# Создание задачи.
class LearningTaskCreateView(LoginRequiredMixin, GPermissionRequiredMixin, CreateView):
    # Права доступа.
    permission_required = 'learning_path.add_learningpath'
    # Форма.
    form_class = LearningTaskForm
    # Модель.
    model = LearningTask
    # Шаблон.
    template_name = 'learning_task_edit.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_path = LearningPath.objects.get(pk=self.kwargs.get('pk'))
        return learning_path

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем позицию.
        learning_path = LearningPath.objects.get(pk=self.kwargs.get('pk'))
        last_position = learning_path.learning_tasks.order_by('position').values_list('position', flat=True).last()
        if last_position:
            position = last_position + 1
        else:
            position = 1
        initial["position"] = position
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Добавляем траекторию.
        initial["learning_path"] = learning_path
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_path', kwargs={'pk': self.object.learning_path.pk})

# Сортировка задач траектории.
@login_required
@permission_required('testing.change_learningpath')
def learning_tasks_ordering(request, pk):

    #Кверисет.
    queryset = LearningTask.objects.filter(learning_path__pk=pk).order_by('position')

    # Формсет.
    LearningTaskSet = modelformset_factory(
        LearningTask,
        fields=('type',),
        can_order=True,
        can_delete=False,
        extra=0,
        widgets={
            'type': forms.HiddenInput(),
        },
    )

    # Если форма отправлена.
    if request.method == 'POST':
        formset = LearningTaskSet(request.POST, queryset=queryset)
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data:
                    learning_tasks = form.save(commit=False)
                    learning_tasks.position = form.cleaned_data[ORDERING_FIELD_NAME]
                    learning_tasks.save()
            return redirect('learning_path:learning_path', pk=pk)

    # Если форма открыта.
    else:
        formset = LearningTaskSet(queryset=queryset)

    # Контекст.
    contex = {'formset': formset}

    # Форма.
    return render(request, 'learning_tasks_ordering.html', contex)

# Изменение задачи.
class LearningTaskUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    # Права доступа.
    permission_required = 'learning_path.change_learningpath'
    # Форма.
    form_class = LearningTaskForm
    accept_global_perms = True
    # Модель.
    model = LearningTask
    # Шаблон.
    template_name = 'learning_task_edit.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_path = self.get_object().learning_path
        return learning_path

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:learning_path', kwargs={'pk': self.object.learning_path.pk})

# Удаление задач.
class LearningTaskDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    # Права доступа.
    permission_required = 'learning_path.delete_learningpath'
    accept_global_perms = True
    # Модель.
    model = LearningTask
    # Шаблон.
    template_name = 'learning_task_delete.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        learning_path = self.get_object().learning_path
        return learning_path

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объектов.
        return reverse('learning_path:learning_path', kwargs={'pk': self.object.learning_path.pk})

# Список назначений.
class AssignmentsView(LoginRequiredMixin, PreviousPageSetMixinL1, PermissionListMixin, ListView):
    # Права доступа
    permission_required = 'learning_path.view_assignment'
    # Модель.
    model = Assignment
    # Поле сортировки.
    ordering = 'learning_path__name'
    # Шаблон.
    template_name = 'assignments.html'
    # Количество объектов на странице
    paginate_by = 6

    # Переопределяем выборку вью.
    def get_queryset(self):
        # Забираем изначальную выборку вью.
        queryset = super().get_queryset()
        # Добавляем модель фильтрации в выборку вью.
        self.filterset = AssignmentFilter(self.request.GET, queryset, request=self.request)
        # Возвращаем вью новую выборку.
        return self.filterset.qs

    # Добавляем во вью переменные.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)
        # Добавляем фильтрсет.
        context['filterset'] = self.filterset
        # Возвращаем новый набор переменных.
        return context

# Объект назначения.
class AssignmentView(LoginRequiredMixin, PreviousPageGetMixinL1, PermissionRequiredMixin, DetailView):
    # Права доступа.
    permission_required = 'learning_path.view_assignment'
    accept_global_perms = True
    # Модель.
    model = Assignment
    # Шаблон.
    template_name = 'assignment.html'

    # Добавляем во вью переменные.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)

        # Добавляем повторы.
        if self.object.repeats.exists():
            repeats = self.object.repeats.all().order_by('-created')
            repeats_count = repeats.count()
        else:
            repeats = AssignmentRepeat.objects.none()
            repeats_count = 0

        # Добавляем пагинатор.
        repeats_paginator = Paginator(repeats, 6)
        repeats_page_number = self.request.GET.get('repeats_page')
        repeats_page_obj = repeats_paginator.get_page(repeats_page_number)
        # Добавляем во вью.
        context['repeats_page_obj'] = repeats_page_obj
        context['repeats_count'] = repeats_count

        # Возвращаем новый набор переменных.
        return context

# Создание задачи.
class AssignmentCreateView(LoginRequiredMixin, GPermissionRequiredMixin, CreateView):
    # Права доступа.
    permission_required = 'learning_path.add_assignment'
    # Форма.
    form_class = AssignmentForm
    # Модель.
    model = Assignment
    # Шаблон.
    template_name = 'assignment_edit.html'

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:assignment', kwargs={'pk': self.object.pk})

# Изменение задачи.
class AssignmentUpdateView(LoginRequiredMixin, GPermissionRequiredMixin, UpdateView):
    # Права доступа.
    permission_required = 'learning_path.add_assignment'
    accept_global_perms = True
    # Форма.
    form_class = AssignmentForm
    # Модель.
    model = Assignment
    # Шаблон.
    template_name = 'assignment_edit.html'

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        initial["planned_start_date"] = self.object.planned_start_date.strftime('%Y-%m-%d')
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:assignment', kwargs={'pk': self.object.pk})

# Удаление назначения.
class AssignmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    # Права доступа.
    permission_required = 'learning_path.delete_assignment'
    accept_global_perms = True
    # Модель.
    model = Assignment
    # Шаблон.
    template_name = 'assignment_delete.html'

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объектов.
        return reverse('learning_path:assignments')

# Создание задачи.
class AssignmentRepeatCreateView(LoginRequiredMixin, GPermissionRequiredMixin, CreateView):
    # Права доступа.
    permission_required = 'learning_path.add_assignment'
    # Форма.
    form_class = AssignmentRepeatForm
    # Модель.
    model = AssignmentRepeat
    # Шаблон.
    template_name = 'assignment_repeat_edit.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        assignment = Assignment.objects.get(pk=self.kwargs.get('pk'))
        return assignment

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Добавляем назначение.
        assignment = self.get_permission_object()
        initial["assignment"] = assignment
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:assignment', kwargs={'pk': self.object.assignment.pk})

# Изменение задачи.
class AssignmentRepeatUpdateView(LoginRequiredMixin, GPermissionRequiredMixin, UpdateView):
    # Права доступа.
    permission_required = 'learning_path.add_assignment'
    accept_global_perms = True
    # Форма.
    form_class = AssignmentRepeatForm
    # Модель.
    model = AssignmentRepeat
    # Шаблон.
    template_name = 'assignment_repeat_edit.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        assignment = self.get_object().assignment
        return assignment

    # Заполнение полей данными.
    def get_initial(self):
        # Забираем изначальный набор.
        initial = super().get_initial()
        # Добавляем создателя: юзера отправившего запрос.
        initial["creator"] = self.request.user
        # Добавляем назначение.
        assignment = self.get_permission_object()
        initial["assignment"] = assignment
        # Возвращаем значения в форму.
        return initial

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объекта.
        return reverse('learning_path:assignment', kwargs={'pk': self.object.assignment.pk})

# Удаление назначения.
class AssignmentRepeatDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    # Права доступа.
    permission_required = 'learning_path.delete_assignment'
    accept_global_perms = True
    # Модель.
    model = AssignmentRepeat
    # Шаблон.
    template_name = 'assignment_repeat_delete.html'

    # Определяем объект проверки.
    def get_permission_object(self):
        assignment = self.get_object().assignment
        return assignment

    # Перенаправление после валидации формы.
    def get_success_url(self):
        # Направляем по адресу объектов.
        return reverse('learning_path:assignment', kwargs={'pk': self.object.assignment.pk})

# Самоназначение.
@login_required
def self_appointment(request, pk, type):

    appoint = True

    # Если траектория.
    if type == 'learning_path':

        # Забираем.
        learning_path = LearningPath.objects.get(pk=pk)

        # Если нельзя назначить.
        if not learning_path.self_appointment:
            return HttpResponseForbidden('Эту траекторию нельзя назначить самому!')

        # Если результаты есть.
        if Result.objects.filter(employee=request.user, type='learning_path', learning_path=learning_path).exists():
            last_result = Result.objects.filter(employee=request.user, type='learning_path', learning_path=learning_path).latest('id')

            # Если уже идет.
            if last_result.status == 'appointed' or last_result.status == 'in_progress':
                appoint = False
                if settings.DEBUG:
                    logger.info(f"Существующий последний результат: {last_result}")
                return HttpResponseForbidden('Вы уже обучаетесь! Перейдите в меню.')

        # Если результата нет.
        if appoint == True:

            # Создание объекта результата для учебной траектории.
            learning_path_result = Result.objects.create(
                learning_path=learning_path,
                employee=request.user,
                type='learning_path',
                self_appointment = True
            )
            logger.info(f"Создан {learning_path_result}")

            # Перебор задач учебного пути.
            for learning_task in learning_path.learning_tasks.all():
                logger.info(f"Назначение учебной задачи {learning_task} для {request.user}.")

                if learning_task.type == 'material':

                    # Забираем материал.
                    material = learning_task.material

                    # Создание объекта результата для задачи.
                    learning_task_result = Result.objects.create(
                        learning_path=learning_path,
                        learning_path_result=learning_path_result,
                        employee=request.user,
                        material=material,
                        learning_task=learning_task,
                        type='material',
                        self_appointment=True
                    )
                    logger.info(f"Создан {learning_task_result}")

                if learning_task.type == 'test':

                    # Забираем тест.
                    test = learning_task.test

                    # Создание объекта результата для задачи.
                    learning_task_result = Result.objects.create(
                        learning_path=learning_path,
                        learning_path_result=learning_path_result,
                        employee=request.user,
                        test=test,
                        learning_task=learning_task,
                        type='test',
                        self_appointment=True
                    )
                    logger.info(f"Создан {learning_task_result}")

                if learning_task.type == 'course':

                    # Забираем курс.
                    course = learning_task.course

                    # Создание объекта результата для задачи.
                    learning_task_result = Result.objects.create(
                        learning_path=learning_path,
                        learning_path_result=learning_path_result,
                        employee=request.user,
                        course=course,
                        scorm_package=course.scorm_package,
                        learning_task=learning_task,
                        type='course',
                        self_appointment=True
                    )
                    logger.info(f"Создан {learning_task_result}")

            # Уходим.
            return redirect('learning_path:learning_path', pk=pk)

    # Если материал.
    if type == 'material':

        # Забираем.
        material = Material.objects.get(pk=pk)

        # Если нельзя назначить.
        if not material.self_appointment:
            return HttpResponseForbidden('Этот материал нельзя назначить самому!')

        # Если результаты есть.
        if Result.objects.filter(employee=request.user, material=material).exists():
            last_result = Result.objects.filter(employee=request.user, material=material).latest('id')

            # Если уже идет.
            if last_result.status == 'appointed' or last_result.status == 'in_progress':
                appoint = False
                if settings.DEBUG:
                    logger.info(f"Существующий последний результат: {last_result}")
                return HttpResponseForbidden('Вы уже обучаетесь! Перейдите в меню.')

        # Если результата нет.
        if appoint == True:

            # Создаем.
            result = Result.objects.create(
                employee=request.user,
                type = 'material',
                material=material,
                self_appointment = True
            )
            logger.info(f"Создан результат: {result}")

            # Уходим.
            return redirect('materials:material', pk=pk)

    # Если курс.
    if type == 'course':

        # Забираем.
        course = Course.objects.get(pk=pk)

        # Если нельзя назначить.
        if not course.self_appointment:
            return HttpResponseForbidden('Этот курс нельзя назначить самому!')

        # Если уже идет.
        if Result.objects.filter(employee=request.user, course=course).exists():
            last_result = Result.objects.filter(employee=request.user, course=course).latest('id')

            # Если уже пройден не переназначаем.
            if last_result.status == 'appointed' or last_result.status == 'in_progress':
                appoint = False
                if settings.DEBUG:
                    logger.info(f"Существующий последний результат: {last_result}")
                return HttpResponseForbidden('Вы уже обучаетесь! Перейдите в меню.')

        # Если результата нет.
        if appoint == True:

            # Создаем.
            result = Result.objects.create(
                employee=request.user,
                type='course',
                course=course,
                scorm_package=course.scorm_package,
                self_appointment=True
            )
            logger.info(f"Создан результат: {result}")

            # Уходим.
            return redirect('courses:course', pk=pk)

    # Если тест.
    if type == 'test':

        # Забираем.
        test = Test.objects.get(pk=pk)

        # Если нельзя назначить.
        if not test.self_appointment:
            return HttpResponseForbidden('Этот тест нельзя назначить самому!')

        # Если уже идет.
        if Result.objects.filter(employee=request.user, test=test).exists():
            last_result = Result.objects.filter(employee=request.user, test=test).latest('id')

            # Если уже пройден не переназначаем.
            if last_result.status == 'appointed' or last_result.status == 'in_progress':
                appoint = False
                if settings.DEBUG:
                    logger.info(f"Существующий последний результат: {last_result}")
                return HttpResponseForbidden('Вы уже обучаетесь! Перейдите в меню.')

        # Если результата нет.
        if appoint == True:

            # Создаем.
            result = Result.objects.create(
                employee=request.user,
                type='test',
                test=test,
                self_appointment=True
            )
            logger.info(f"Создан результат: {result}")

            # Уходим.
            return redirect('testing:test', pk=pk)

    # Если эвент.
    if type == 'event':

        # Забираем.
        event = Event.objects.get(pk=pk)

        # Если нельзя назначить.
        if not event.self_appointment:
            return HttpResponseForbidden('Это мероприятие нельзя назначить самому!')

        # Если результаты есть.
        if Result.objects.filter(employee=request.user, event=event).exists():
            last_result = Result.objects.filter(employee=request.user, event=event).latest('id')

            # Если уже идет.
            if last_result.status == 'appointed' or last_result.status == 'in_progress':
                appoint = False
                if settings.DEBUG:
                    logger.info(f"Существующий последний результат: {last_result}")
                return HttpResponseForbidden('Вы уже обучаетесь! Перейдите в меню.')

        # Если результата нет.
        if appoint == True:

            # Создаем.
            result = Result.objects.create(
                employee=request.user,
                type='event',
                event=event,
                status='registered',
                planned_end_date=event.date,
                self_appointment=True
            )
            logger.info(f"Создан результат: {result}")

            # Уходим.
            return redirect('events:event', pk=pk)


# Список траекторий.
class ResultsView(LoginRequiredMixin, PreviousPageSetMixinL1, ListView):
    # Модель.
    model = Result
    # Поле сортировки.
    ordering = '-planned_end_date'
    # Шаблон.
    template_name = 'results.html'
    # Количество объектов на странице
    paginate_by = 6

    # Проверяем права.
    def dispatch(self, request, *args, **kwargs):

        # Если имеется полное разрешение.
        if request.user.has_perm('learning_path.view_result'):
            # Идем дальше.
            return super().dispatch(request, *args, **kwargs)

        # Если имеется разрешение руководителя.
        if request.user.has_perm('learning_path.view_managed_result'):

            # Получаем корневые подразделения, для которых пользователь является менеджером.
            sub_subdivisions = Subdivision.objects.filter(
                positions__placements__manager=True,
                positions__placements__employee=request.user
            ).distinct().prefetch_related('subdivisions')

            # Забираем переменные.
            sub_subdivisions_list = list(sub_subdivisions)
            n = 0

            # Пока не пройдены все подчиненные подразделения.
            while n < len(sub_subdivisions_list):
                # Забираем текущее.
                current_sub_subdivision = sub_subdivisions_list[n]
                # Проходим циклом по их дочерним подразделелениям.
                for sub_subdivision in current_sub_subdivision.subdivisions.all():
                    # Добавляем их дочерние подразделения в список своих подчиненных.
                    if sub_subdivision not in sub_subdivisions_list:
                        sub_subdivisions_list.append(sub_subdivision)
                # Переходим к следующему.
                n += 1

            # Фильтруем результаты запроса по управляемым подразделениям.
            self.queryset = Result.objects.filter(
                employee__placements__position__subdivision__in=sub_subdivisions
            ).distinct()

            # Идем дальше.
            return super().dispatch(request, *args, **kwargs)

        # Если нет ни каких прав.
        else:

            # Запрет.
            return HttpResponseForbidden('У Вас нет прав на просмотр результатов!')

    # Переопределяем метод get
    def get(self, request, *args, **kwargs):

        # Получаем параметры фильтрации из GET-запроса
        filter_params = request.GET.copy()

        # Если параметр export_excel есть, удаляем его из запроса и генерируем отчет Excel.
        if 'export_excel' in filter_params:
            filter_params.pop('export_excel')

            # Получаем фильтрсет с этими параметрами.
            filterset = ResultFilter(filter_params, queryset=Result.objects.all())

            # Применяем фильтрсет к queryset и получаем отфильтрованные результаты.
            filtered_results = filterset.qs

            # Создаем файл Excel с помощью openpyxl.
            output = io.BytesIO()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Результаты ' + datetime.now().strftime('%d.%m.%Y %H-%M')

            # Добавляем заголовки столбцов.
            headers = ['Логин', 'ФИО', 'Организация', 'Подразделение', 'Должность', 'Назначение', 'Комплексная программа', 'Траектория обучения', 'Учебная задача', 'Планируемая дата завершения', 'Дата завершения', 'Статус', 'Полученный балл в %']
            worksheet.append(headers)

            # Записываем данные в строки.
            for result in filtered_results:
                organization_names = list(result.employee.placements.all().values_list('position__subdivision__organization__legal_name', flat=True))
                subdivision_names = list(result.employee.placements.all().values_list('position__subdivision__name', flat=True))
                position_names = list(result.employee.placements.all().values_list('position__name', flat=True))
                row = [
                    result.employee.username,
                    f'{result.employee.last_name} {result.employee.first_name} {result.employee.fathers_name}'
                    if result.employee.last_name or result.employee.first_name or result.employee.fathers_name else '',
                    ", ".join(organization_names),
                    ", ".join(subdivision_names),
                    ", ".join(position_names),
                    str(result.assignment) if result.assignment else
                    ('Cамоназначение' if result.self_appointment else
                     ('Приглашен')),
                    str(result.learning_complex) if result.learning_complex and not result.learning_path else '',
                    str(result.learning_path) if result.learning_path and not result.learning_task else '',
                    str(result.learning_task) if result.learning_task else
                    (str(result.material) if result.material else
                     (str(result.course) if result.course else
                      (str(result.test) if result.test else
                       str(result.event) if result.event else ''))),
                    result.planned_end_date.strftime('%d.%m.%Y') if result.planned_end_date else None,
                    result.end_date.strftime('%d.%m.%Y %H:%M') if result.end_date else None,
                    result.get_status_display() if not result.event else result.get_status_display(),
                    str(result.score_scaled) if result.score_scaled else '',
                ]

                worksheet.append(row)

            # Задаем ширину столбцов.
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 30
            worksheet.column_dimensions['D'].width = 30
            worksheet.column_dimensions['E'].width = 30
            worksheet.column_dimensions['F'].width = 30
            worksheet.column_dimensions['G'].width = 30
            worksheet.column_dimensions['H'].width = 30
            worksheet.column_dimensions['I'].width = 30
            worksheet.column_dimensions['J'].width = 30
            worksheet.column_dimensions['K'].width = 30

            # Задаем перенос строк в ячейках.
            for row in worksheet.iter_rows(min_row=2, max_row=len(filtered_results) + 1):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            # Задаем границы.
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            for row in worksheet.iter_rows(min_row=1, max_row=len(filtered_results) + 1, min_col=1, max_col=13):
                for cell in row:
                    cell.border = thin_border

            # Сохраняем файл Excel.
            workbook.save(output)
            output.seek(0)

            # Отправляем файл Excel в HTTP-ответ.
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Results_' + datetime.now().strftime('%d.%m.%Y %H-%M') + '.xlsx'
            return response

        # Иначе продолжаем обработку запроса как обычно
        return super().get(request, *args, **kwargs)

    # Переопределяем выборку вью.
    def get_queryset(self):
        # Забираем изначальную выборку вью.
        queryset = super().get_queryset()
        # Добавляем модель фильтрации в выборку вью.
        self.filterset = ResultFilter(self.request.GET, queryset, request=self.request)
        # Возвращаем вью новую выборку.
        return self.filterset.qs

    # Добавляем во вью переменные.
    def get_context_data(self, **kwargs):
        # Забираем изначальный набор переменных.
        context = super().get_context_data(**kwargs)
        # Добавляем фильтрсет.
        context['filterset'] = self.filterset
        # Возвращаем новый набор переменных.
        return context